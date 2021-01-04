from openpyxl import load_workbook
from openpyxl.styles import Font


def transaction():
    # Loads workbook and different sheets
    workbook = load_workbook(filename="Accounting.xlsx")
    sheet = workbook["Journal"]
    ledger = workbook["Ledger"]

    # Set Row, this uses the last row used and adds 2 to it
    count = sheet.max_row + 2
    # Set Journal Account Columns
    Date = "A"
    Day = "B"
    Acc = "C"
    Dr = "E"
    Cr = "F"

    # Set Ledger Accounts

    Bank = "C1"
    Cash = "I1"
    AR = "O1"
    AP = "U1"

    # Will need Special treatment when dealing with
    Inc = "AA1"
    Exp = "AG1"

    print("Please enter DD-MM-YYYY: ")
    date = str(input())

    day = date[0] + date[1]
    month_year = f'{date[2] + date[3]}/{date[4] + date[5] + date[6] + date[7]}'

    sheet[Date + str(count)] = month_year
    count += 1
    sheet[Day + str(count)] = day

    dr = int(input("How many Debit accounts: "))
    dr_amount = []
    debit_acc = []
    i = 0
    print("Enter Debit Account: ")
    while i < dr:
        acc = str(input())
        debit_acc.append(acc)
        sheet[Acc + str(count)] = debit_acc[i]
        amount = float(input("Enter amount: "))
        dr_amount.append(amount)
        sheet[Dr + str(count)] = amount
        count += 1
        i += 1
        if dr > 1 and i < dr:
            print("Enter Debit Account: ")

    cr = int(input("How many Credit accounts: "))
    credit_acc = []
    k = 0
    print("Enter Credit Account: ")
    while k < cr:
        acc = str(input())
        credit_acc.append(acc)
        sheet[Acc + str(count)] = credit_acc[k]
        sheet[Acc + str(count)].font = Font(name='Arial', size=10, bold=True)
        amount = float(input("Enter amount: "))
        sheet[Cr + str(count)] = amount
        count += 1
        k += 1
        if cr > 1 and k < cr:
            print("Enter Credit Account: ")
    descr = str(input("Enter Transaction Description: "))
    sheet[Acc + str(count)] = descr
    sheet[Acc + str(count)].font = Font(name='Arial', size=10, italic=True)

    # First we will iterate through debit accounts
    l = 0
    while l < len(debit_acc):
        if debit_acc[l] == "Bank":
            col = ord(Bank[0])
            row = len(ledger['A']) + 2

            ledger[chr(col - 2) + str(row)] = f'{day}/{month_year}'  # Date
            ledger[chr(col - 1) + str(row)] = descr  # Description
            ledger[chr(col) + str(row)] = dr_amount[l]  # Amount
            ledger[chr(col + 2) + str(row)] = f'={chr(col + 2) + str(row - 1)} + {chr(col) + str(row)} - {chr(col + 1) + str(row)}' # Balance
            break
        else:
            l += 1
    workbook.save(filename="Accounting.xlsx")
